from openpyxl import load_workbook
import os
stus = {}
def getfiles(path,suffix):                   #suffix为str，示例:   '.txt'      '.json'       '.h5'
    res = []
    for file in os.listdir("./%s/"%path):
        name,suf = os.path.splitext(file)
        if suf == suffix:
            res.append(file)
    return(res)
def get_filter_files(dir, ext=None):
    allfiles = []
    needExtFilter = (ext != None)
    for root, dirs, files in os.walk(dir):
        for filepath in files:
            filepath = os.path.join(root, filepath)
            extension = os.path.splitext(filepath)[1][1:]
            if needExtFilter and extension in ext:
                allfiles.append(filepath)
            elif not needExtFilter:
                allfiles.append(filepath)
    return allfiles
def get_dirs():
    res = []
    for file in os.listdir("."):
        if os.path.isdir(file):
            fs=getfiles(file,".xlsx")
            if len(fs) == 1:
                res.append(file)
    return(res)
def deal_dir(path):
    fs = get_filter_files("./%s/考勤/"%path,["xlsx"])
    for f in fs:
        w = load_workbook(f)
        s1=w.worksheets[0]
        for r in s1.rows:
            if r[6].value == "缺勤":
                print(r[0].value,r[6].value)
                if r[0].value not in stus:
                    stus[r[0].value] = [0,0,0,0]
                stus[r[0].value][0] +=1
            elif r[6].value == "迟到":
                print(r[0].value,r[6].value)
                if r[0].value not in stus:
                    stus[r[0].value] = [0,0,0,0]
                stus[r[0].value][1] +=1
    print("考勤完成")
    print(stus)
    fs = get_filter_files("./%s/作业/"%path,["xlsx"])
    for f in fs:
        if "一课一文" in f:
            continue
        w = load_workbook(f)
        s1=w.worksheets[0]
        for r in s1.rows:
            if r[8].value == "0":
                print(r[1].value,r[8].value)
                if r[1].value not in stus:
                    stus[r[1].value] = [0,0,0,0]
                stus[r[1].value][2] +=1
    print("作业完成")
    print(stus)
    w = load_workbook("./%s/作业/一课一文.xlsx"%path)
    s1=w.worksheets[0]
    for r in s1.rows:
        if r[2].value == "1971":
            print(r[1].value,r[8].value)
            if r[1].value not in stus:
                stus[r[1].value] = [0,0,0,0]
            stus[r[1].value][3] = int(r[8].value)
    print("一课一文完成")
    print(stus)
    fs=getfiles(path,".xlsx")
    if len(fs) != 1:
        print("无点名册")
        exit(0)
    w = load_workbook("%s/%s"%(path,fs[0]))
    s1=w.worksheets[0]
    for r in s1.rows:
        if str(r[0].value).isdigit():
            if r[3].value in stus:
                if stus[r[3].value][0]>0:
                    s1.merge_cells("F%d:L%d"%(r[0].row,r[0].row))
                    s1["F%d"%r[0].row].value = "缺勤%d次"%stus[r[3].value][0]
                if stus[r[3].value][1]>0:
                    s1.merge_cells("M%d:R%d"%(r[0].row,r[0].row))
                    s1["M%d"%r[0].row].value = "迟到%d次"%stus[r[3].value][1]
                r[20].value = 20 - stus[r[3].value][0]*5 - stus[r[3].value][1] *2
                if r[20].value < 0:
                    r[20].value = 0
                r[27].value = 20 - stus[r[3].value][2]*5
                if r[27].value < 0:
                    r[27].value = 0
                r[32].value = stus[r[3].value][3]
    w.save("new_%s"%fs[0])
for p in get_dirs():
    deal_dir(p)
print("全部完成")